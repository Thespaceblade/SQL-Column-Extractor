#!/usr/bin/env python3
"""
Extract table.column references from SQL files and output to CSV/Excel format.

EASIEST WAY TO USE:
    1. Open this file and find the FOLDER_PATH variable near the top
    2. Paste your folder path there (e.g., FOLDER_PATH = "/path/to/your/sql/files")
    3. Run: python extract_columns.py

ALTERNATIVE USAGE:
    python extract_columns.py                    # Extract from all SQL files in current directory
    python extract_columns.py 01_simple_ctes.sql  # Extract from specific file
    python extract_columns.py folder/ --output columns.csv  # Process folder, output to CSV
    python extract_columns.py --output columns.xlsx  # Output to Excel
"""

import sys
import csv
import re
import html
import urllib.parse
import logging
import shutil
from pathlib import Path
from typing import Optional
from collections import defaultdict
from datetime import datetime

try:
    import sqlglot
    from sqlglot import expressions as exp
    from sqlglot.errors import ParseError
except ImportError:
    print("Error: sqlglot is not installed.")
    print("Install with: pip install sqlglot")
    print("Or for better performance: pip install 'sqlglot[rs]'")
    sys.exit(1)

try:
    import sqlparse
    SQLPARSE_AVAILABLE = True
except ImportError:
    SQLPARSE_AVAILABLE = False
    print("Warning: sqlparse is not installed. Fallback parsing will be disabled.")
    print("Install with: pip install sqlparse")


# ============================================================================
# CONFIGURATION: Paste your folder path here
# ============================================================================
# Paste the full path to your SQL files folder here, or leave empty to use
# command-line arguments or process current directory
# Examples:
#   FOLDER_PATH = "/Users/username/Documents/SQL_Queries"
#   FOLDER_PATH = r"C:\Users\username\Documents\SQL_Queries"  # Windows
#   FOLDER_PATH = ""  # Leave empty to use command-line args or current directory
# ============================================================================
FOLDER_PATH = r""


def is_rdl_skip_file(filepath: Path) -> bool:
    """
    Check if file should be skipped based on RDL-aware filename patterns.
    
    Skips files matching patterns like:
    - SOR_DATA, SORDATA
    - SOR_DATA_REFRESH, SORDATAREFRESH
    - SORREFRESH
    - Any variation with underscores or without
    
    Args:
        filepath: Path object to the SQL file
        
    Returns:
        bool: True if file should be skipped
    """
    filename = filepath.name.upper()
    
    skip_patterns = [
        r'SOR[_\s]*DATA',
        r'SOR[_\s]*DATA[_\s]*REFRESH',
        r'SOR[_\s]*REFRESH',
        r'DATA[_\s]*REFRESH',
        r'REFRESH[_\s]*DATA',
    ]
    
    for pattern in skip_patterns:
        if re.search(pattern, filename):
            return True
    
    return False


def parse_filename(filepath: Path) -> tuple[str, str]:
    """
    Parse filename to extract Report Name and Dataset.
    
    Format: <report_name>__<dataset>.sql
    If no double underscore, Dataset defaults to "Default"
    
    Args:
        filepath: Path object to the SQL file
        
    Returns:
        tuple: (report_name, dataset)
    """
    filename = filepath.name
    
    if filename.lower().endswith('.sql'):
        filename = filename[:-4]
    
    if '__' in filename:
        parts = filename.split('__', 1)
        report_name = parts[0]
        dataset = parts[1] if len(parts) > 1 else "Default"
    else:
        report_name = filename
        dataset = "Default"
    
    return report_name, dataset


def should_skip_dataset(dataset: str) -> bool:
    """
    Check if a dataset should be skipped based on name patterns.
    
    Skips datasets containing:
    - SOR% (e.g., SOR_DATA, SOR_REFRESH, SORDATA, etc.)
    - Tablix
    - EndDate
    - EvidenceTab
    - EvidenceTablix
    
    Args:
        dataset: Dataset name to check
        
    Returns:
        bool: True if dataset should be skipped
    """
    dataset_upper = dataset.upper()
    
    # Check for SOR% pattern (case-insensitive)
    if 'SOR' in dataset_upper:
        return True
    
    # Check for Tablix (case-insensitive)
    if 'TABLIX' in dataset_upper:
        return True
    
    # Check for EndDate (case-insensitive)
    if 'ENDDATE' in dataset_upper:
        return True
    
    # Check for EvidenceTab (case-insensitive)
    if 'EVIDENCETAB' in dataset_upper:
        return True
    
    # Check for EvidenceTablix (case-insensitive)
    if 'EVIDENCETABLIX' in dataset_upper:
        return True
    
    return False


def decode_html_entities(sql: str) -> str:
    """
    Convert HTML encoded characters to their actual symbols.
    
    Handles:
    - HTML entities: &gt; -> >, &lt; -> <, &amp; -> &, &quot; -> ", &apos; -> ', etc.
    - Numeric entities: &#60; -> <, &#62; -> >, &#38; -> &, etc.
    - Hex entities: &#x3C; -> <, &#x3E; -> >, &#x26; -> &, etc.
    - URL-encoded characters: %3E -> >, %3C -> <, %26 -> &, etc.
    
    Args:
        sql: SQL string that may contain HTML/URL encoded characters
        
    Returns:
        SQL string with HTML/URL encoded characters decoded
    """
    sql = html.unescape(sql)
    
    # Handle double-encoded URL entities (decode multiple times if needed)
    max_decode_iterations = 3
    for _ in range(max_decode_iterations):
        try:
            decoded = urllib.parse.unquote(sql)
            if decoded == sql:
                break
            sql = decoded
        except Exception:
            break
    
    # NOTE: Do NOT replace '+' signs globally - they are valid SQL operators
    # URL-encoded plus signs (%2B) are already handled by urllib.parse.unquote above
    
    html_entity_map = {
        '&gt;': '>',
        '&lt;': '<',
        '&amp;': '&',
        '&quot;': '"',
        '&apos;': "'",
        '&nbsp;': ' ',
        '&copy;': '(c)',
        '&reg;': '(R)',
        '&trade;': '(TM)',
        '&#39;': "'",
        '&#x27;': "'",
        '&#x22;': '"',
        '&#x26;': '&',
        '&#x3C;': '<',
        '&#x3E;': '>',
        '&#x60;': '`',
        '&#x7B;': '{',
        '&#x7D;': '}',
        '&#x7C;': '|',
        '&#60;': '<',
        '&#62;': '>',
        '&#38;': '&',
        '&#34;': '"',
        '&#96;': '`',
        '&#123;': '{',
        '&#125;': '}',
        '&#124;': '|',
    }
    
    for entity, char in html_entity_map.items():
        sql = sql.replace(entity, char)
    
    # Handle any remaining numeric/hex HTML entities (catch-all pattern)
    sql = re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)) if int(m.group(1), 16) < 0x110000 else m.group(0), sql)
    sql = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))) if int(m.group(1)) < 0x110000 else m.group(0), sql)
    
    return sql


def preprocess_sql(sql: str) -> str:
    """
    Preprocess SQL to handle edge cases and remove problematic syntax.
    
    Handles:
    - SQL comments (-- and /* */)
    - DDL statements (CREATE, ALTER, DROP)
    - DECLARE and SET statements
    - WITH (NOLOCK) hints
    - USE statements
    - Isolation levels
    - SET NOCOUNT ON/OFF
    - ASCII control characters
    - Escape codes
    - HTML entities
    - TOP clause (removed, but columns still extracted)
    - Non-printable characters
    - Normalize whitespace
    """
    sql = decode_html_entities(sql)
    
    # Handle BOM variants (UTF-8, UTF-16 BE/LE)
    if sql.startswith('\ufeff'):
        sql = sql[1:]
    if sql.startswith('\xFE\xFF') or sql.startswith('\xFF\xFE'):
        sql = sql[2:]
    
    # Normalize line endings to LF (handle CRLF, CR, LF)
    sql = sql.replace('\r\n', '\n').replace('\r', '\n')
    
    sql = re.sub(r'--.*?$', '', sql, flags=re.MULTILINE)
    sql = re.sub(r'/\*.*?\*/', '', sql, flags=re.DOTALL)
    sql = re.sub(r'(?i)(?:^|[\n;])\s*USE\s+[^\s;]+(?:\s*;)?\s*(?=[\n;]|$|\s)', '', sql, flags=re.MULTILINE)
    sql = re.sub(r'(?i)(?:^|[\n;]|\s)\s*GO\s+\d+\s*(?=[\n;]|$|\s)', '', sql, flags=re.MULTILINE)
    sql = re.sub(r'(?i)(?:^|[\n;]|\s)\s*GO\s*;?\s*(?=[\n;]|$|\s)', '', sql, flags=re.MULTILINE)
    sql = re.sub(r'(?i)SET\s+NOCOUNT\s+(ON|OFF)\s*;?', '', sql, flags=re.MULTILINE)
    sql = re.sub(r'(?i)SET\s+TRANSACTION\s+ISOLATION\s+LEVEL\s+[^;]+;?', '', sql, flags=re.MULTILINE)
    
    # Remove transaction control statements
    sql = re.sub(r'(?i)\bBEGIN\s+TRANSACTION\s*;?', '', sql)
    sql = re.sub(r'(?i)\bCOMMIT\s+(TRANSACTION)?\s*;?', '', sql)
    sql = re.sub(r'(?i)\bROLLBACK\s+(TRANSACTION)?\s*;?', '', sql)
    sql = re.sub(r'(?i)\bSAVE\s+TRANSACTION\s+\w+\s*;?', '', sql)
    
    # Remove TRY/CATCH blocks
    sql = re.sub(r'(?i)\bBEGIN\s+TRY\s*;?', '', sql)
    sql = re.sub(r'(?i)\bBEGIN\s+CATCH\s*;?', '', sql)
    sql = re.sub(r'(?i)\bEND\s+(TRY|CATCH)\s*;?', '', sql)
    
    statements = sql.split(';')
    cleaned_statements = []
    for stmt in statements:
        stmt = stmt.strip()
        if not stmt:
            continue
        if re.match(r'(?i)^\s*SET\s+', stmt) and not re.search(r'(?i)\bUPDATE\s+.*\bSET\b', stmt):
            continue
        cleaned_statements.append(stmt)
    sql = '; '.join(cleaned_statements)
    
    sql = re.sub(r'(?i)\bDECLARE\s+[^;]+;?', '', sql)
    sql = re.sub(r'(?i)\b(CREATE|ALTER|DROP)\s+[^;]+;?', '', sql)
    
    # Remove SQL Server hints (comprehensive list)
    hints_pattern = r'(?i)\s+WITH\s*\(\s*(NOLOCK|READPAST|READUNCOMMITTED|READCOMMITTED|REPEATABLEREAD|SERIALIZABLE|UPDLOCK|XLOCK|ROWLOCK|PAGLOCK|TABLOCK|TABLOCKX|FASTFIRSTROW|FORCESEEK|FORCESCAN|NOWAIT|READCOMMITTEDLOCK|READPASTLOCK)\s*\)'
    sql = re.sub(hints_pattern, '', sql)
    sql = re.sub(r'(?i)\(\s*NOLOCK\s*\)', '', sql)
    sql = re.sub(r'(?i)\s+\(NOLOCK\)', '', sql)
    sql = re.sub(r'(?i)\bTOP\s*\(\s*\d+\s*\)\s+', '', sql)
    sql = re.sub(r'(?i)\bTOP\s+\d+\s+', '', sql)
    
    ansi_csi = re.compile(r'\x9B\[[0-9;]*[a-zA-Z@-~]')
    sql = ansi_csi.sub('', sql)
    sql = re.sub(r'\x9B[^\x20-\x7E]*', '', sql)
    
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[/]*[@-~])')
    sql = ansi_escape.sub('', sql)
    ansi_escape_octal = re.compile(r'\033(?:[@-Z\\-_]|\[[0-?]*[/]*[@-~])')
    sql = ansi_escape_octal.sub('', sql)
    
    sql = re.sub(r'\x1B\][^\x07\x1B]*(\x07|\x1B\\)', '', sql)
    sql = re.sub(r'\033\][^\x07\x1B]*(\x07|\x1B\\)', '', sql)
    sql = re.sub(r'\x1B P[^\x1B]*\x1B\\', '', sql)
    sql = re.sub(r'\033 P[^\x1B]*\x1B\\', '', sql)
    sql = re.sub(r'\x1B\^[^\x1B]*\x1B\\', '', sql)
    sql = re.sub(r'\033\^[^\x1B]*\x1B\\', '', sql)
    sql = re.sub(r'\x1B_[^\x1B]*\x1B\\', '', sql)
    sql = re.sub(r'\033_[^\x1B]*\x1B\\', '', sql)
    
    # Remove standalone escape characters (ESC and CSI) - do this AFTER removing complete sequences
    # This won't affect SQL brackets [], quotes "", or other valid SQL syntax
    sql = re.sub(r'[\x1B\x9B]', '', sql)
    sql = re.sub(r'\\x1[bB]', '', sql)
    sql = re.sub(r'\\033', '', sql)
    sql = re.sub(r'\\x9[bB]', '', sql)
    sql = re.sub(r'\\x[0-9a-fA-F]{2}', '', sql)
    sql = re.sub(r'\\[0-7]{1,3}', '', sql)
    sql = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', sql)
    sql = re.sub(r'[\u200B-\u200F\uFEFF\u2060\u2061-\u2064\u2028\u2029\u2066-\u2069\u061C\u2000-\u200A]', '', sql)
    sql = re.sub(r'[\u00AD\u034F\u180E]', '', sql)
    
    # Remove other Unicode control characters
    sql = re.sub(r'[\u202A-\u202E\u206A-\u206F\uFFF0-\uFFFF]', '', sql)
    
    # Remove non-breaking spaces and other Unicode spaces
    sql = re.sub(r'[\u00A0\u2000-\u200B\u202F\u205F\u3000]', ' ', sql)
    
    # Normalize whitespace
    sql = re.sub(r'[ \t]+', ' ', sql)
    sql = re.sub(r'\n{2,}', '\n', sql)
    sql = re.sub(r'[ \t]+$', '', sql, flags=re.MULTILINE)
    sql = sql.strip()
    
    return sql


def strip_brackets(identifier):
    """
    Strip SQL Server brackets [] from an identifier.
    
    Args:
        identifier: String or identifier object that may contain brackets
        
    Returns:
        String with brackets removed, or original value if not a string
    """
    if not identifier:
        return identifier
    
    identifier_str = identifier if isinstance(identifier, str) else str(identifier)
    
    # Remove brackets: [name] -> name
    if identifier_str.startswith('[') and identifier_str.endswith(']'):
        return identifier_str[1:-1]
    
    return identifier_str


def build_scope_alias_maps(stmt: exp.Expression) -> dict:
    """
    Build scope-aware alias maps where each SELECT scope has its own alias map.
    Returns dict mapping select_node_id -> alias_map for that scope.
    Also returns a global CTE names set.
    
    This handles alias shadowing where the same alias is used in different scopes.
    """
    scope_maps = {}  # id(select_node) -> {alias -> table}
    cte_names = set()
    
    # First pass: collect all CTE names globally
    for cte in stmt.find_all(exp.CTE):
        if cte.alias:
            cte_name = cte.alias if isinstance(cte.alias, str) else cte.alias.name if hasattr(cte.alias, 'name') else str(cte.alias)
            if cte_name:
                cte_names.add(strip_brackets(cte_name) if isinstance(cte_name, str) else cte_name)
    
    def extract_table_alias_for_scope(table_expr, local_alias_map, context_cte_names=None):
        """Extract table name and alias from a table expression into local scope map."""
        if context_cte_names is None:
            context_cte_names = set()
        
        if isinstance(table_expr, exp.Table):
            table_name = table_expr.name
            if not table_name:
                return
            
            # Strip brackets from catalog, db, and table name before building qualified name
            parts = []
            if table_expr.catalog:
                catalog_clean = strip_brackets(table_expr.catalog)
                parts.append(catalog_clean)
            if table_expr.db:
                db_clean = strip_brackets(table_expr.db)
                parts.append(db_clean)
            if table_name:
                table_name_clean = strip_brackets(table_name)
                parts.append(table_name_clean)
            
            full_table_name = ".".join(parts) if parts else table_name_clean
            alias = table_expr.alias
            alias_name = None
            if alias:
                if isinstance(alias, exp.TableAlias):
                    alias_this = alias.this
                    if isinstance(alias_this, exp.Identifier):
                        alias_name = alias_this.name
                    elif isinstance(alias_this, str):
                        alias_name = alias_this
                    else:
                        alias_name = str(alias_this)
                elif isinstance(alias, str):
                    alias_name = alias
                else:
                    alias_name = str(alias)
            
            # Use cleaned table name for CTE check
            if table_name_clean in cte_names or table_name_clean in context_cte_names:
                resolved_name = table_name_clean
            else:
                resolved_name = full_table_name
            
            if alias_name:
                # Strip brackets from alias name before storing
                alias_name_clean = strip_brackets(alias_name)
                local_alias_map[alias_name_clean] = resolved_name
                local_alias_map[alias_name_clean.lower()] = resolved_name
                if alias_name != alias_name_clean:
                    local_alias_map[alias_name] = resolved_name
                    if isinstance(alias_name, str):
                        local_alias_map[alias_name.lower()] = resolved_name
            
            # Store table name as well
            local_alias_map[table_name_clean] = resolved_name
            if isinstance(table_name_clean, str):
                local_alias_map[table_name_clean.lower()] = resolved_name
            if table_name != table_name_clean:
                local_alias_map[table_name] = resolved_name
                if isinstance(table_name, str):
                    local_alias_map[table_name.lower()] = resolved_name
        
        elif isinstance(table_expr, (exp.Subquery, exp.Lateral)):
            # Handle subqueries, derived tables, CROSS APPLY, OUTER APPLY
            alias = table_expr.alias
            if alias:
                if isinstance(alias, exp.TableAlias):
                    alias_this = alias.this
                    if isinstance(alias_this, exp.Identifier):
                        alias_name = alias_this.name
                    elif isinstance(alias_this, str):
                        alias_name = alias_this
                    else:
                        alias_name = str(alias_this)
                elif isinstance(alias, str):
                    alias_name = alias
                else:
                    alias_name = str(alias)
                
                if alias_name:
                    alias_name_clean = strip_brackets(alias_name)
                    local_alias_map[alias_name_clean] = alias_name_clean
                    local_alias_map[alias_name_clean.lower()] = alias_name_clean
    
    def build_scope_map(select_node, parent_scope_id=None):
        """Build alias map for a single SELECT scope, inheriting from parent."""
        # Start with copy of parent scope's aliases (inheritance)
        if parent_scope_id and parent_scope_id in scope_maps:
            local_map = dict(scope_maps[parent_scope_id])
        else:
            local_map = {}
        
        # Add CTE names
        for cte_name in cte_names:
            local_map[cte_name] = cte_name
            if isinstance(cte_name, str):
                local_map[cte_name.lower()] = cte_name
        
        # Extract tables from FROM clause
        from_expr = select_node.args.get("from") or select_node.args.get("from_")
        if from_expr:
            if isinstance(from_expr, exp.From):
                extract_table_alias_for_scope(from_expr.this, local_map, cte_names)
            elif isinstance(from_expr, exp.Table):
                extract_table_alias_for_scope(from_expr, local_map, cte_names)
        
        # Extract tables from JOINs
        for join in select_node.args.get("joins", []):
            if isinstance(join, exp.Join):
                extract_table_alias_for_scope(join.this, local_map, cte_names)
        
        # Store this scope's map
        scope_maps[id(select_node)] = local_map
        return id(select_node)
    
    def process_expression(expr, parent_scope_id=None):
        """Recursively process expression tree building scope maps."""
        if isinstance(expr, exp.Select):
            # Build scope map for this SELECT
            current_scope_id = build_scope_map(expr, parent_scope_id)
            # Process nested expressions with current scope as parent
            for child in expr.walk():
                if child != expr:
                    if isinstance(child, exp.Select):
                        process_expression(child, current_scope_id)
        elif isinstance(expr, (exp.Union, exp.Except, exp.Intersect)):
            if expr.this:
                process_expression(expr.this, parent_scope_id)
            if expr.expression:
                process_expression(expr.expression, parent_scope_id)
        elif isinstance(expr, exp.CTE):
            if expr.this:
                process_expression(expr.this, parent_scope_id)
    
    # Process the main statement
    if isinstance(stmt, exp.Select):
        process_expression(stmt)
    else:
        # Process all SELECT statements in order
        for select in stmt.find_all(exp.Select):
            # Only process top-level SELECTs (others handled recursively)
            parent = select.parent
            is_top_level = not isinstance(parent, (exp.Select, exp.Subquery))
            if is_top_level or id(select) not in scope_maps:
                process_expression(select)
    
    return scope_maps, cte_names


def get_scope_for_column(col: exp.Column, scope_maps: dict) -> dict:
    """
    Find the appropriate scope (alias map) for a given column by walking up the AST.
    Returns the alias map for the closest containing SELECT.
    """
    node = col
    while node:
        if isinstance(node, exp.Select) and id(node) in scope_maps:
            return scope_maps[id(node)]
        node = node.parent
    # Return empty dict if no scope found
    return {}


def build_alias_map(stmt: exp.Expression) -> dict:
    """
    Build a mapping of table aliases to actual table/CTE names.
    Returns dict mapping alias -> full_table_name or cte_name
    Handles all cases: FROM aliases, JOIN aliases, CTE aliases, subquery aliases
    
    NOTE: This function builds a FLAT alias map (last definition wins).
    For scope-aware resolution, use build_scope_alias_maps() instead.
    
    Strips brackets from schema and table names to ensure clean alias mapping.
    """
    alias_map = {}
    cte_names = set()
    
    for cte in stmt.find_all(exp.CTE):
        if cte.alias:
            cte_name = cte.alias if isinstance(cte.alias, str) else cte.alias.name if hasattr(cte.alias, 'name') else str(cte.alias)
            if cte_name:
                cte_names.add(cte_name)
                alias_map[cte_name] = cte_name
                if isinstance(cte_name, str):
                    alias_map[cte_name.lower()] = cte_name
    
    def extract_table_alias(table_expr, alias_map, cte_names, context_cte_names=None):
        """Extract table name and alias from a table expression."""
        if context_cte_names is None:
            context_cte_names = set()
        
        if isinstance(table_expr, exp.Table):
            table_name = table_expr.name
            if not table_name:
                return
            
            # Strip brackets from catalog, db, and table name before building qualified name
            parts = []
            if table_expr.catalog:
                catalog_clean = strip_brackets(table_expr.catalog)
                parts.append(catalog_clean)
            if table_expr.db:
                db_clean = strip_brackets(table_expr.db)
                parts.append(db_clean)
            if table_name:
                table_name_clean = strip_brackets(table_name)
                parts.append(table_name_clean)
            
            full_table_name = ".".join(parts) if parts else table_name_clean
            alias = table_expr.alias
            alias_name = None
            if alias:
                if isinstance(alias, exp.TableAlias):
                    alias_this = alias.this
                    if isinstance(alias_this, exp.Identifier):
                        alias_name = alias_this.name
                    elif isinstance(alias_this, str):
                        alias_name = alias_this
                    else:
                        alias_name = str(alias_this)
                elif isinstance(alias, str):
                    alias_name = alias
                else:
                    alias_name = str(alias)
            
            # Use cleaned table name for CTE check
            if table_name_clean in cte_names or table_name_clean in context_cte_names:
                resolved_name = table_name_clean
            else:
                resolved_name = full_table_name
            
            if alias_name:
                # Strip brackets from alias name before storing (to match how we look it up later)
                alias_name_clean = strip_brackets(alias_name)
                alias_map[alias_name_clean] = resolved_name
                alias_map[alias_name_clean.lower()] = resolved_name
                # Also store original alias (with brackets) for lookup compatibility
                if alias_name != alias_name_clean:
                    alias_map[alias_name] = resolved_name
                    if isinstance(alias_name, str):
                        alias_map[alias_name.lower()] = resolved_name
            
            # Store both original and cleaned table name in alias map
            alias_map[table_name_clean] = resolved_name
            if isinstance(table_name_clean, str):
                alias_map[table_name_clean.lower()] = resolved_name
            # Also store original table_name (with brackets) mapping to cleaned resolved name
            if table_name != table_name_clean:
                alias_map[table_name] = resolved_name
                if isinstance(table_name, str):
                    alias_map[table_name.lower()] = resolved_name
        
        elif isinstance(table_expr, (exp.Subquery, exp.Lateral)):
            # Handle subqueries, derived tables, CROSS APPLY, OUTER APPLY
            alias = table_expr.alias
            if alias:
                if isinstance(alias, exp.TableAlias):
                    alias_this = alias.this
                    if isinstance(alias_this, exp.Identifier):
                        alias_name = alias_this.name
                    elif isinstance(alias_this, str):
                        alias_name = alias_this
                    else:
                        alias_name = str(alias_this)
                elif isinstance(alias, str):
                    alias_name = alias
                else:
                    alias_name = str(alias)
                
                # Store the derived table/subquery alias (maps to itself since it's a derived table)
                if alias_name:
                    alias_name_clean = strip_brackets(alias_name)
                    alias_map[alias_name_clean] = alias_name_clean
                    alias_map[alias_name_clean.lower()] = alias_name_clean
    
    def extract_from_joins(expression, context_cte_names=None):
        """Recursively extract table aliases from FROM and JOIN clauses."""
        if context_cte_names is None:
            context_cte_names = set()
        
        local_cte_names = set(context_cte_names)
        for cte in expression.find_all(exp.CTE):
            if cte.alias:
                cte_name = cte.alias if isinstance(cte.alias, str) else cte.alias.name if hasattr(cte.alias, 'name') else str(cte.alias)
                if cte_name:
                    local_cte_names.add(cte_name)
        
        if isinstance(expression, exp.Select):
            from_expr = expression.args.get("from") or expression.args.get("from_")
            if from_expr:
                if isinstance(from_expr, exp.From):
                    extract_table_alias(from_expr.this, alias_map, cte_names, local_cte_names)
                elif isinstance(from_expr, exp.Table):
                    extract_table_alias(from_expr, alias_map, cte_names, local_cte_names)
            
            for join in expression.args.get("joins", []):
                if isinstance(join, exp.Join):
                    extract_table_alias(join.this, alias_map, cte_names, local_cte_names)
        
        if isinstance(expression, (exp.Union, exp.Except, exp.Intersect)):
            if expression.this:
                extract_from_joins(expression.this, local_cte_names)
            if expression.expression:
                extract_from_joins(expression.expression, local_cte_names)
        
        for cte in expression.find_all(exp.CTE):
            if cte.this:
                extract_from_joins(cte.this, local_cte_names)
        
        for subquery in expression.find_all(exp.Subquery):
            if subquery.this:
                extract_from_joins(subquery.this, local_cte_names)
        
        for select_stmt in expression.find_all(exp.Select):
            if select_stmt != expression:
                extract_from_joins(select_stmt, local_cte_names)
    
    extract_from_joins(stmt)
    return alias_map


def resolve_unqualified_columns(stmt: exp.Expression, alias_map: dict) -> dict:
    """
    Build a mapping of unqualified column names to their source tables.
    Returns dict mapping column_name -> table_name for unqualified columns.
    Now handles WHERE clauses and multiple tables by checking column context.
    """
    column_to_table = {}
    
    case_insensitive_alias_map = {}
    for key, value in alias_map.items():
        case_insensitive_alias_map[key.lower()] = value
        case_insensitive_alias_map[key] = value
    
    def get_tables_from_from_clause(select_stmt):
        """Get all table names from FROM and JOIN clauses."""
        tables = []
        
        from_expr = select_stmt.args.get("from") or select_stmt.args.get("from_")
        if from_expr:
            if isinstance(from_expr, exp.From):
                table_expr = from_expr.this
                if isinstance(table_expr, exp.Table):
                    table_name = table_expr.name if isinstance(table_expr.name, str) else str(table_expr.name)
                    if table_name:
                        alias = table_expr.alias
                        if alias:
                            if isinstance(alias, exp.TableAlias):
                                alias_name = alias.this.name if isinstance(alias.this, exp.Identifier) else str(alias.this)
                            else:
                                alias_name = str(alias)
                            # Strip brackets from alias name before lookup (to match how it's stored)
                            alias_name_clean = strip_brackets(alias_name)
                            resolved = case_insensitive_alias_map.get(alias_name_clean.lower()) or case_insensitive_alias_map.get(alias_name_clean) or alias_map.get(alias_name_clean) or alias_map.get(alias_name)
                            if resolved:
                                tables.append(resolved)
                            else:
                                tables.append(table_name)
                        else:
                            resolved = case_insensitive_alias_map.get(table_name.lower()) or alias_map.get(table_name)
                            if resolved:
                                tables.append(resolved)
                            else:
                                tables.append(table_name)
        
        for join in select_stmt.args.get("joins", []):
            if isinstance(join, exp.Join):
                table_expr = join.this
                if isinstance(table_expr, exp.Table):
                    table_name = table_expr.name if isinstance(table_expr.name, str) else str(table_expr.name)
                    if table_name:
                        alias = table_expr.alias
                        if alias:
                            if isinstance(alias, exp.TableAlias):
                                alias_name = alias.this.name if isinstance(alias.this, exp.Identifier) else str(alias.this)
                            else:
                                alias_name = str(alias)
                            # Strip brackets from alias name before lookup (to match how it's stored)
                            alias_name_clean = strip_brackets(alias_name)
                            resolved = case_insensitive_alias_map.get(alias_name_clean.lower()) or case_insensitive_alias_map.get(alias_name_clean) or alias_map.get(alias_name_clean) or alias_map.get(alias_name)
                            if resolved:
                                tables.append(resolved)
                            else:
                                tables.append(table_name)
                        else:
                            resolved = case_insensitive_alias_map.get(table_name.lower()) or alias_map.get(table_name)
                            if resolved:
                                tables.append(resolved)
                            else:
                                tables.append(table_name)
        
        return tables
    
    def find_column_source_table(column_name, select_stmt, alias_map, case_insensitive_alias_map):
        """
        Try to determine which table a column belongs to by checking:
        1. If there's exactly one table, use it
        2. If column appears in JOIN conditions, use the table from that JOIN
        3. If column appears in WHERE with table prefix in same expression, use that table
        """
        tables = get_tables_from_from_clause(select_stmt)
        
        if len(tables) == 1:
            return tables[0]
        
        for join in select_stmt.args.get("joins", []):
            if isinstance(join, exp.Join):
                join_condition = join.args.get("on") or join.args.get("using")
                if join_condition:
                    for col in join_condition.find_all(exp.Column):
                        if col.name:
                            col_name = col.name if isinstance(col.name, str) else str(col.name)
                            if col_name.lower() == column_name.lower():
                                if col.table:
                                    table_ref = col.table if isinstance(col.table, str) else str(col.table)
                                    resolved = case_insensitive_alias_map.get(table_ref.lower()) or alias_map.get(table_ref)
                                    if resolved:
                                        return resolved
                                    return table_ref
        
        where_clause = select_stmt.args.get("where")
        if where_clause:
            for col in where_clause.find_all(exp.Column):
                if col.name:
                    col_name = col.name if isinstance(col.name, str) else str(col.name)
                    if col_name.lower() == column_name.lower() and col.table:
                        table_ref = col.table if isinstance(col.table, str) else str(col.table)
                        resolved = case_insensitive_alias_map.get(table_ref.lower()) or alias_map.get(table_ref)
                        if resolved:
                            return resolved
                        return table_ref
        
        return None
    
    # Process all SELECT statements
    for select_stmt in stmt.find_all(exp.Select):
        # Find all unqualified columns in this SELECT (including WHERE, JOIN, HAVING, etc.)
        for col in select_stmt.find_all(exp.Column):
            if col.name and not col.table:
                column_name = col.name if isinstance(col.name, str) else str(col.name)
                # Try to find source table using improved logic
                source_table = find_column_source_table(column_name, select_stmt, alias_map, case_insensitive_alias_map)
                if source_table:
                    # Use case-insensitive key for lookup
                    column_to_table[column_name.lower()] = source_table
                    column_to_table[column_name] = source_table  # Keep original case too
    
    return column_to_table


def format_parse_error(error: Exception, sql: str, dialect: Optional[str] = None, statement_num: Optional[int] = None) -> str:
    """
    Format a parse error with detailed information.
    
    Args:
        error: The parse error exception
        sql: The SQL string that failed to parse
        dialect: The dialect that was being used
        statement_num: Statement number if processing multiple statements
        
    Returns:
        Formatted error message string
    """
    error_msg = []
    
    # Error type and basic message
    error_type = type(error).__name__
    error_msg.append(f"Parse Error ({error_type})")
    
    if statement_num is not None:
        error_msg.append(f"Statement #{statement_num + 1}")
    
    if dialect:
        error_msg.append(f"Dialect: {dialect}")
    else:
        error_msg.append("Dialect: Generic SQL")
    
    error_msg.append("")
    
    # Get error message - ParseError has better structure
    line_num = None
    col_num = None
    
    if isinstance(error, ParseError):
        # ParseError has an 'errors' attribute with structured error info
        if hasattr(error, 'errors') and error.errors:
            # Use the first error's description (most relevant)
            first_error = error.errors[0]
            error_description = first_error.get('description', str(error))
            # Clean ANSI codes from error description
            error_description = re.sub(r'\x1b\[[0-9;]*m', '', error_description)
            error_description = re.sub(r'\033\[[0-9;]*m', '', error_description)
            error_description = re.sub(r'\[[0-9;]*m', '', error_description)
            error_msg.append(f"Error: {error_description}")
            
            # Extract line and column from error structure
            line_num = first_error.get('line')
            col_num = first_error.get('col')
            
            if line_num:
                error_msg.append(f"Line: {line_num}")
            if col_num:
                error_msg.append(f"Column: {col_num}")
        else:
            # Fallback to string representation
            error_str = str(error)
            # Remove ANSI codes from error message if present
            error_str = re.sub(r'\x1b\[[0-9;]*m', '', error_str)
            error_str = re.sub(r'\033\[[0-9;]*m', '', error_str)
            error_str = re.sub(r'\[[0-9;]*m', '', error_str)
            error_msg.append(f"Error: {error_str}")
            
            # Try to find line number in error message
            line_match = re.search(r'line\s+(\d+)', error_str, re.IGNORECASE)
            col_match = re.search(r'col(umn)?\s+(\d+)', error_str, re.IGNORECASE)
            
            if line_match:
                line_num = int(line_match.group(1))
            if col_match:
                col_num = int(col_match.group(2))
    else:
        # For non-ParseError exceptions
        error_str = str(error)
        # Remove ANSI codes from error message if present
        error_str = re.sub(r'\x1b\[[0-9;]*m', '', error_str)
        error_str = re.sub(r'\033\[[0-9;]*m', '', error_str)
        error_str = re.sub(r'\[[0-9;]*m', '', error_str)
        error_msg.append(f"Error: {error_str}")
        
        # Try to find line number in error message
        line_match = re.search(r'line\s+(\d+)', error_str, re.IGNORECASE)
        col_match = re.search(r'col(umn)?\s+(\d+)', error_str, re.IGNORECASE)
        
        if line_match:
            line_num = int(line_match.group(1))
        if col_match:
            col_num = int(col_match.group(2))
    
    # Show context around error line if we have line number
    if line_num:
        sql_lines = sql.split('\n')
        if 1 <= line_num <= len(sql_lines):
            start_line = max(0, line_num - 3)
            end_line = min(len(sql_lines), line_num + 2)
            
            error_msg.append("\nContext:")
            for i in range(start_line, end_line):
                line_prefix = ">>> " if i == line_num - 1 else "    "
                error_msg.append(f"{line_prefix}{i+1:4d}: {sql_lines[i]}")
    
    # Common error suggestions
    error_msg.append("\nSuggestions:")
    
    # Get error string for suggestions (use description if available, otherwise full error)
    error_str_for_suggestions = ""
    if isinstance(error, ParseError) and hasattr(error, 'errors') and error.errors:
        error_str_for_suggestions = error.errors[0].get('description', str(error)).lower()
    else:
        error_str_for_suggestions = str(error).lower()
    
    if "unexpected" in error_str_for_suggestions or "syntax" in error_str_for_suggestions:
        error_msg.append("  - Check for missing commas, parentheses, or quotes")
        error_msg.append("  - Verify SQL syntax matches the specified dialect")
        error_msg.append("  - Check for unclosed quotes or parentheses")
    
    if "unknown" in error_str_for_suggestions or "invalid" in error_str_for_suggestions:
        error_msg.append("  - Verify table/column names are correct")
        error_msg.append("  - Check for reserved keywords that need quoting")
        error_msg.append("  - Ensure dialect-specific syntax is correct")
    
    if not dialect:
        error_msg.append("  - Try specifying a dialect: --dialect tsql|mssql|postgres|mysql|snowflake|oracle|bigquery")
        error_msg.append("    (Note: 'tsql'/'mssql' is for SQL Server - 'mssql' is automatically converted to 'tsql')")
    
    return "\n".join(error_msg)


def normalize_dialect(dialect: Optional[str]) -> Optional[str]:
    """
    Normalize dialect names to sqlglot-compatible names.
    Converts common aliases to their sqlglot equivalents.
    Defaults to 'tsql' if None.
    
    Args:
        dialect: Dialect name (may be None or an alias)
    
    Returns:
        Normalized dialect name (defaults to 'tsql')
    """
    if not dialect:
        return 'tsql'
    
    dialect_lower = dialect.lower().strip()
    
    # Map common aliases to sqlglot dialect names
    dialect_map = {
        'mssql': 'tsql',
        'sqlserver': 'tsql',
        'sql_server': 'tsql',
        'sql-server': 'tsql',
        't-sql': 'tsql',
        'tsql': 'tsql',  # Already correct
        'postgres': 'postgres',
        'postgresql': 'postgres',
        'mysql': 'mysql',
        'snowflake': 'snowflake',
        'oracle': 'oracle',
        'bigquery': 'bigquery',
        'big_query': 'bigquery',
    }
    
    return dialect_map.get(dialect_lower, dialect)  # Return original if not in map


def fallback_extract_columns(sql: str, enable_unqualified_resolution: bool = True) -> list:
    """
    Fallback tolerant parser using sqlparse tokenizer + regex extraction.
    Used when AST parsing fails or yields zero columns.
    
    Args:
        sql: SQL string to parse
        enable_unqualified_resolution: Whether to infer table names for unqualified columns
    
    Returns:
        List of table.column references found via regex
    """
    if not SQLPARSE_AVAILABLE:
        return []
    
    columns = []
    
    try:
        # Tokenize SQL using sqlparse (tolerant, doesn't require valid SQL)
        parsed = sqlparse.parse(sql)
        
        if not parsed:
            return []
        
        sql_text = sql.upper()
        
        # Extract table aliases using regex patterns
        # Pattern: FROM table [AS] alias or JOIN table [AS] alias
        alias_patterns = [
            r'(?:FROM|JOIN)\s+(\w+)(?:\s+(?:AS\s+)?(\w+))?',
            r'(\w+)\s+(?:AS\s+)?(\w+)\s+(?:JOIN|,|WHERE)',
        ]
        
        alias_map = {}
        for pattern in alias_patterns:
            for match in re.finditer(pattern, sql_text, re.IGNORECASE):
                table = match.group(1)
                alias = match.group(2) if match.lastindex >= 2 and match.group(2) else None
                if alias and alias.upper() not in ['AS', 'ON', 'WHERE', 'AND', 'OR']:
                    alias_map[alias.upper()] = table.upper()
                # Also map table to itself
                alias_map[table.upper()] = table.upper()
        
        # Extract table.column patterns using regex
        # Pattern: table.column or alias.column
        column_pattern = r'\b(\w+)\.(\w+)\b'
        
        for match in re.finditer(column_pattern, sql_text):
            table_ref = match.group(1).upper()
            col_name = match.group(2).upper()
            
            # Skip if it's a function call (e.g., COUNT(*), MAX(col))
            if col_name == '*':
                continue
            
            # Resolve alias to table name
            actual_table = alias_map.get(table_ref, table_ref)
            
            # Build qualified name
            qualified_name = f"{actual_table}.{col_name}"
            columns.append(qualified_name)
        
        # If unqualified resolution is enabled, try to infer table names
        if enable_unqualified_resolution and columns:
            # Find unqualified columns that appear after FROM/JOIN
            # This is a simple heuristic - find column names that aren't table.column
            unqualified_pattern = r'\b(?<!\.)(\w+)(?!\.)\b'
            # This is simplified - in practice, you'd want more context
            pass
        
    except Exception:
        # If fallback also fails, return empty list
        pass
    
    return columns


def extract_table_columns(sql: str, dialect: Optional[str] = None, filepath: Optional[str] = None, error_details: Optional[list] = None, enable_unqualified_resolution: bool = True, try_multiple_dialects: bool = False) -> tuple[list, str]:
    """
    Extract all table.column references from SQL.
    Resolves aliases to full table names and qualifies unqualified columns.
    Returns ALL occurrences (not unique) - every time a table.column is referenced.
    
    Args:
        sql: SQL string to parse
        dialect: SQL dialect to use (None = auto-detect)
        filepath: Optional filepath for better error messages
    
    Returns:
        List of strings in format "table.column" or "schema.table.column"
        Only includes columns that have a table reference (no standalone columns or tables)
    """
    columns = []
    
    # Normalize dialect name (convert mssql -> tsql, etc.) - defaults to 'tsql'
    dialect = normalize_dialect(dialect)
    
    # Determine dialects to try
    if try_multiple_dialects:
        dialects_to_try = [dialect] if dialect else ['tsql', 'postgres', 'mysql', 'snowflake', 'oracle', 'bigquery']
    else:
        # Default: only try the specified dialect (or tsql if None)
        dialects_to_try = [dialect] if dialect else ['tsql']
    
    last_error = None
    last_dialect = None
    failed_statements = []
    dialects_tried = []  # Track which dialects we actually attempted
    ast_succeeded = False
    parse_error_occurred = False
    
    for try_dialect in dialects_to_try:
        dialect_name = try_dialect or 'tsql'  # Default to tsql instead of generic
        dialects_tried.append(dialect_name)
        
        try:
            statements = sqlglot.parse(sql, dialect=try_dialect)
            
            # Check if parsing succeeded (at least one non-None statement)
            if not statements or all(s is None for s in statements):
                # All statements failed to parse
                if len(statements) > 0:
                    # We got statements but they're all None - parsing failed
                    parse_error_msg = f"Failed to parse SQL with dialect '{try_dialect or 'tsql'}': All statements returned None"
                    last_error = ParseError(parse_error_msg)
                    last_dialect = try_dialect
                    parse_error_occurred = True
                    
                    # Capture this error in error_details
                    if error_details is not None:
                        formatted_error = format_parse_error(last_error, sql, try_dialect)
                        error_details.append({
                            'statement': 'all',
                            'dialect': try_dialect or 'tsql',
                            'error': parse_error_msg,
                            'formatted': formatted_error
                        })
                # If trying multiple dialects, continue; otherwise break to try fallback
                if try_multiple_dialects:
                    continue
                else:
                    break
            
            # Process each statement individually, skip ones that fail
            for i, stmt in enumerate(statements):
                if stmt is None:
                    failed_statements.append((i, try_dialect, "Statement parsed as None"))
                    # Capture None statement errors
                    if error_details is not None:
                        error_msg = f"Statement {i+1} failed to parse (returned None) with dialect '{try_dialect or 'tsql'}'"
                        # Try to create a formatted error
                        try:
                            # Create a simple error for None statements
                            none_error = ParseError(error_msg)
                            formatted_error = format_parse_error(none_error, sql, try_dialect, i)
                        except:
                            formatted_error = f"Parse Error\nStatement #{i+1}\nDialect: {try_dialect or 'tsql'}\n\nError: {error_msg}"
                        
                        error_details.append({
                            'statement': i + 1,
                            'dialect': try_dialect or 'tsql',
                            'error': error_msg,
                            'formatted': formatted_error
                        })
                    continue
                
                try:
                    # Build scope-aware alias maps for this statement
                    scope_maps, global_cte_names = build_scope_alias_maps(stmt)
                    
                    # Also build flat alias map for fallback (last definition wins)
                    alias_map = build_alias_map(stmt)
                    
                    # Create case-insensitive alias map
                    case_insensitive_alias_map = {}
                    for key, value in alias_map.items():
                        case_insensitive_alias_map[key.lower()] = value
                    
                    # Resolve unqualified columns to their source tables (if enabled)
                    unqualified_map = {}
                    if enable_unqualified_resolution:
                        unqualified_map = resolve_unqualified_columns(stmt, alias_map)
                    
                    # Helper function to get fallback table from FROM clause
                    def get_fallback_table(select_stmt):
                        """Get the first table from FROM clause as fallback for unresolvable unqualified columns."""
                        from_expr = select_stmt.args.get("from") or select_stmt.args.get("from_")
                        if from_expr:
                            if isinstance(from_expr, exp.From):
                                table_expr = from_expr.this
                            else:
                                table_expr = from_expr
                            
                            if isinstance(table_expr, exp.Table):
                                parts = []
                                if table_expr.catalog:
                                    parts.append(strip_brackets(table_expr.catalog if isinstance(table_expr.catalog, str) else str(table_expr.catalog)))
                                if table_expr.db:
                                    parts.append(strip_brackets(table_expr.db if isinstance(table_expr.db, str) else str(table_expr.db)))
                                if table_expr.name:
                                    parts.append(strip_brackets(table_expr.name if isinstance(table_expr.name, str) else str(table_expr.name)))
                                
                                if parts:
                                    # Resolve alias if present
                                    alias = table_expr.alias
                                    if alias:
                                        alias_name = alias.this.name if isinstance(alias, exp.TableAlias) and isinstance(alias.this, exp.Identifier) else str(alias)
                                        # Strip brackets from alias name before lookup (to match how it's stored)
                                        alias_name_clean = strip_brackets(alias_name)
                                        resolved = alias_map.get(alias_name_clean) or alias_map.get(alias_name_clean.lower()) or alias_map.get(alias_name) if alias_name else None
                                        if resolved:
                                            return resolved
                                    
                                    # Check if table name is in alias map
                                    table_name_only = parts[-1]
                                    resolved = alias_map.get(table_name_only) or alias_map.get(table_name_only.lower())
                                    if resolved:
                                        return resolved
                                    
                                    return ".".join(parts)
                        return None
                    
                    # Find all column references (including WHERE, JOIN, HAVING, etc.)
                    for col in stmt.find_all(exp.Column):
                        if col.name:
                            # Track if this column was originally unqualified (no table reference)
                            was_unqualified = col.table is None or (isinstance(col.table, str) and not col.table.strip())
                            
                            # Get table name (either from column or from unqualified mapping)
                            # Strip brackets from table name
                            table_name = col.table
                            if table_name:
                                table_name = strip_brackets(table_name if isinstance(table_name, str) else str(table_name))
                            
                            # If column is unqualified, try to resolve it (case-insensitive)
                            if not table_name:
                                col_name = col.name if isinstance(col.name, str) else str(col.name)
                                # Try case-insensitive lookup
                                table_name = (unqualified_map.get(col_name) or 
                                            unqualified_map.get(col_name.lower()))
                            
                            # If still no table name and column was originally unqualified, use fallback
                            if not table_name and was_unqualified:
                                # Try to get fallback table from FROM clause
                                # Find the SELECT statement this column belongs to
                                current_select = None
                                for select in stmt.find_all(exp.Select):
                                    if col in select.find_all(exp.Column):
                                        current_select = select
                                        break
                                
                                if current_select:
                                    fallback_table = get_fallback_table(current_select)
                                    if fallback_table:
                                        table_name = fallback_table
                            
                            # Handle columns without table reference
                            if not table_name:
                                # If column was originally unqualified and we couldn't resolve or find fallback, skip it
                                continue
                            
                            # If column had a table reference originally, verify it can be resolved
                            # (Skip columns with unresolvable table references, but include unqualified columns)
                            
                            # Get scope-specific alias map for this column
                            scope_alias_map = get_scope_for_column(col, scope_maps)
                            # Create case-insensitive version of scope alias map
                            scope_ci_map = {k.lower(): v for k, v in scope_alias_map.items()} if scope_alias_map else {}
                            
                            if not was_unqualified:
                                # Column had a table reference - check if it's resolvable
                                table_name_str = table_name if isinstance(table_name, str) else str(table_name)
                                # Try scope-aware lookup first, then fall back to global
                                resolved_table_check = (scope_ci_map.get(table_name_str.lower()) or
                                                       scope_alias_map.get(table_name_str) or
                                                       case_insensitive_alias_map.get(table_name_str.lower()) or 
                                                       alias_map.get(table_name_str))
                                # If table reference can't be resolved and isn't a direct table name from FROM clause,
                                # skip it (This handles cases like nonexistent_table.column)
                                # Note: We allow it if it's in alias_map as a key (table name) or value (resolved name)
                                if not resolved_table_check:
                                    # Check if it's a direct table name vs unresolved alias
                                    # We need to distinguish between:
                                    #   - A table name that appears as a VALUE in alias_map (valid, was resolved)
                                    #   - A table name that appears as a KEY mapping to itself (direct table)
                                    #   - An alias that appears as KEY but couldn't be resolved (should skip)
                                    
                                    # Check if it's a resolved table name (appears as value in alias_map)
                                    is_resolved_value = any(
                                        table_name_str.lower() == str(v).lower() or 
                                        table_name_str.lower() == str(v).lower().split('.')[-1]
                                        for v in alias_map.values()
                                    )
                                    
                                    # Check if it's a direct table name (appears as key mapping to itself)
                                    is_direct_table = (
                                        (table_name_str in alias_map and alias_map[table_name_str] == table_name_str) or
                                        (table_name_str.lower() in case_insensitive_alias_map and 
                                         case_insensitive_alias_map[table_name_str.lower()].lower() == table_name_str.lower())
                                    )
                                    
                                    # Check if it's embedded in a resolved value (part of schema.table)
                                    is_embedded_table = any(
                                        table_name_str.lower() in str(v).lower().split('.')
                                        for v in alias_map.values()
                                    )
                                    
                                    # Only allow if it's a valid table name, not an unresolved alias
                                    is_known_table = is_resolved_value or is_direct_table or is_embedded_table
                                    
                                    if not is_known_table:
                                        # Table reference is an unresolved alias - skip it
                                        continue
                            
                            # Build fully qualified name
                            parts = []
                            
                            # Handle catalog - check if it's an alias first
                            catalog_part = None
                            if col.catalog:
                                # Strip brackets from catalog before processing
                                catalog_str = strip_brackets(col.catalog if isinstance(col.catalog, str) else str(col.catalog))
                                # Check if catalog is actually an alias - use scope-aware lookup first
                                resolved_catalog = (scope_ci_map.get(catalog_str.lower()) or
                                                   scope_alias_map.get(catalog_str) or
                                                   case_insensitive_alias_map.get(catalog_str.lower()) or 
                                                   alias_map.get(catalog_str))
                                if resolved_catalog:
                                    # It's an alias - resolve it (might be multi-part)
                                    catalog_parts = resolved_catalog.split('.')
                                    parts.extend(catalog_parts)
                                    catalog_part = catalog_parts[0] if catalog_parts else None
                                else:
                                    # Not an alias, use as-is
                                    parts.append(catalog_str)
                                    catalog_part = catalog_str
                            
                            # Handle database/schema - check if it's an alias first
                            db_part = None
                            db_resolved_to_table = False
                            if col.db:
                                # Strip brackets from db before processing
                                db_str = strip_brackets(col.db if isinstance(col.db, str) else str(col.db))
                                # Check if db is actually an alias - use scope-aware lookup first
                                resolved_db = (scope_ci_map.get(db_str.lower()) or
                                              scope_alias_map.get(db_str) or
                                              case_insensitive_alias_map.get(db_str.lower()) or 
                                              alias_map.get(db_str))
                                if resolved_db:
                                    # It's an alias - resolve it
                                    db_parts = resolved_db.split('.')
                                    # If resolved_db contains multiple parts, it's a full table name
                                    # In this case, col.table might be redundant or incorrect
                                    if len(db_parts) > 1:
                                        db_resolved_to_table = True
                                        # If we already have catalog, check for duplication
                                        if catalog_part and len(db_parts) > 0:
                                            if db_parts[0].lower() == catalog_part.lower():
                                                # Skip catalog part if it matches
                                                parts.extend(db_parts[1:] if len(db_parts) > 1 else [])
                                            else:
                                                parts.extend(db_parts)
                                        else:
                                            parts.extend(db_parts)
                                        db_part = db_parts[-1] if db_parts else None
                                    else:
                                        # Single part, treat as schema
                                        if catalog_part and len(db_parts) > 0:
                                            if db_parts[0].lower() == catalog_part.lower():
                                                parts.extend(db_parts[1:] if len(db_parts) > 1 else [])
                                            else:
                                                parts.extend(db_parts)
                                        else:
                                            parts.extend(db_parts)
                                        db_part = db_parts[-1] if db_parts else None
                                else:
                                    # Not an alias, use as-is
                                    parts.append(db_str)
                                    db_part = db_str
                            
                            # Resolve table alias to actual table name (case-insensitive)
                            # BUT: if db resolved to a full table name, col.table might be redundant
                            table_name_str = table_name if isinstance(table_name, str) else str(table_name)
                            
                            # If db resolved to a table name, check if table_name matches the last part
                            if db_resolved_to_table and table_name_str:
                                # db already contains the table name, skip col.table
                                resolved_table = None
                            else:
                                # Use scope-aware lookup first, then fall back to global
                                resolved_table = (scope_ci_map.get(table_name_str.lower()) or
                                                 scope_alias_map.get(table_name_str) or
                                                 case_insensitive_alias_map.get(table_name_str.lower()) or 
                                                 alias_map.get(table_name_str))
                                
                                # If still not found, try stripping brackets and looking up again
                                if not resolved_table:
                                    table_name_stripped = strip_brackets(table_name_str)
                                    if table_name_stripped != table_name_str:
                                        resolved_table = (scope_ci_map.get(table_name_stripped.lower()) or
                                                         scope_alias_map.get(table_name_stripped) or
                                                         case_insensitive_alias_map.get(table_name_stripped.lower()) or
                                                         alias_map.get(table_name_stripped))
                                
                                # If still not found, try case variations
                                if not resolved_table:
                                    table_name_stripped = strip_brackets(table_name_str)
                                    # Try uppercase
                                    resolved_table = alias_map.get(table_name_stripped.upper())
                                    # Try title case
                                    if not resolved_table:
                                        resolved_table = alias_map.get(table_name_stripped.title())
                            
                            if resolved_table:
                                # Replace alias with actual table name
                                actual_table = resolved_table
                                # Split the actual table name and use its parts
                                table_parts = actual_table.split('.')
                                
                                # Avoid duplicating catalog/schema if already in parts
                                if len(parts) > 0:
                                    # Check if resolved table starts with what we already have
                                    # Compare prefix of resolved table with what's in parts
                                    if len(table_parts) >= len(parts):
                                        # Check if prefix matches
                                        prefix_matches = True
                                        for i, part in enumerate(parts):
                                            if i < len(table_parts) and table_parts[i].lower() != part.lower():
                                                prefix_matches = False
                                                break
                                        
                                        if prefix_matches:
                                            # Prefix matches, only add the remaining parts
                                            parts.extend(table_parts[len(parts):])
                                        else:
                                            # Prefix doesn't match, check if just schema matches
                                            if db_part and len(table_parts) > 1:
                                                if table_parts[0].lower() == db_part.lower():
                                                    # Schema matches, skip it
                                                    parts.extend(table_parts[1:])
                                                else:
                                                    # Different, add all
                                                    parts.extend(table_parts)
                                            else:
                                                # No db_part to compare, add all
                                                parts.extend(table_parts)
                                    else:
                                        # Resolved table is shorter, add all parts
                                        parts.extend(table_parts)
                                else:
                                    # No parts yet, include all parts from resolved table
                                    parts.extend(table_parts)
                            else:
                                # Table name not in alias map - determine if it's a valid table name or unresolved alias
                                # CRITICAL: We must NEVER output an alias that couldn't be resolved
                                
                                if db_resolved_to_table:
                                    # db already contains the table name, skip table_name_str
                                    pass
                                else:
                                    # Strategy 1: Check if table_name_str appears as a RESOLVED VALUE in alias_map
                                    # This indicates it's a valid table name that was resolved from another alias
                                    is_resolved_table_name = any(
                                        table_name_str.lower() == str(v).lower() or 
                                        table_name_str.lower() == str(v).lower().split('.')[-1]
                                        for v in alias_map.values()
                                    )
                                    
                                    # Strategy 2: Check if table_name_str is already in parts (qualified name)
                                    is_table_part_of_qualified = any(
                                        part.lower() == table_name_str.lower() 
                                        for part in parts
                                    ) if parts else False
                                    
                                    # Strategy 3: Check if it looks like a full table name (not short alias)
                                    # Short aliases are typically 1-3 chars; table names are longer or contain dots
                                    looks_like_table_name = (
                                        '.' in table_name_str or  # Already qualified
                                        len(table_name_str) > 3 or  # Longer names less likely to be aliases
                                        table_name_str[0].isupper()  # PascalCase table names
                                    ) if table_name_str else False
                                    
                                    # Strategy 4: Check if it's a known CTE name (CTEs map to themselves)
                                    is_cte_name = (
                                        table_name_str in alias_map and 
                                        alias_map[table_name_str] == table_name_str and
                                        table_name_str in global_cte_names
                                    ) if table_name_str else False
                                    
                                    # Only include if it's a valid table name, CTE, or clearly not an alias
                                    if is_resolved_table_name or is_table_part_of_qualified:
                                        parts.append(table_name_str)
                                    elif is_cte_name:
                                        # CTE reference - include as-is (CTEs are valid "tables")
                                        parts.append(table_name_str)
                                    elif looks_like_table_name and not (len(table_name_str) <= 3 and table_name_str.isalpha()):
                                        # Looks like a real table name, not a short alias
                                        parts.append(table_name_str)
                                    else:
                                        # This is likely an unresolved alias - SKIP this column
                                        continue
                            
                            # Add column name
                            col_name = col.name if isinstance(col.name, str) else str(col.name)
                            col_name = strip_brackets(col_name)
                            parts.append(col_name)
                            
                            # Join with dots: schema.table.column or table.column
                            # Must have at least table.column (2 parts minimum)
                            if len(parts) >= 2:
                                qualified_name = ".".join(parts)
                                # Filter out wildcard columns (table.*)
                                if qualified_name.endswith('.*'):
                                    continue
                                columns.append(qualified_name)
                            # Skip columns without table reference
                
                except Exception as e:
                    # Skip this statement but continue with others
                    failed_statements.append((i, try_dialect, str(e)))
                    formatted_error = format_parse_error(e, sql, try_dialect, i)
                    file_info = f" in {filepath}" if filepath else ""
                    
                    # Store error details if list provided
                    if error_details is not None:
                        error_details.append({
                            'statement': i + 1,
                            'dialect': try_dialect or 'tsql',
                            'error': str(e),
                            'formatted': formatted_error
                        })
                    
                    print(f"\n{'='*80}", file=sys.stderr)
                    print(f"Warning: Failed to process statement {i+1}{file_info}", file=sys.stderr)
                    print(formatted_error, file=sys.stderr)
                    print(f"{'='*80}\n", file=sys.stderr)
                    continue
            
            # If we extracted any columns, return SUCCESS
            if columns:
                if failed_statements:
                    print(f"Note: Successfully extracted columns despite {len(failed_statements)} failed statement(s)", file=sys.stderr)
                # AST parsing succeeded and extracted columns
                return columns, "SUCCESS"
            
            # If we got here, AST parsing succeeded but extracted 0 columns
            # Mark that AST succeeded (no exception thrown) and extracted 0 columns
            ast_succeeded = True
            # Reset failed_statements for next dialect attempt
            failed_statements = []
            
            # If trying multiple dialects, continue; otherwise break to try fallback
            if not try_multiple_dialects:
                break
            
        except ParseError as e:
            last_error = e
            last_dialect = try_dialect
            parse_error_occurred = True
            # Store error for this dialect attempt
            if error_details is not None:
                formatted_error = format_parse_error(e, sql, try_dialect)
                # Clean ANSI codes from error string
                error_str = str(e)
                error_str = re.sub(r'\x1b\[[0-9;]*m', '', error_str)
                error_str = re.sub(r'\033\[[0-9;]*m', '', error_str)
                error_str = re.sub(r'\[[0-9;]*m', '', error_str)
                error_details.append({
                    'statement': 'all',
                    'dialect': try_dialect or 'tsql',
                    'error': error_str,
                    'formatted': formatted_error
                })
            # Try next dialect if enabled
            if try_multiple_dialects:
                continue
            else:
                break  # Stop trying dialects
        except Exception as e:
            last_error = e
            last_dialect = try_dialect
            parse_error_occurred = True
            # Store error for this dialect attempt
            if error_details is not None:
                formatted_error = format_parse_error(e, sql, try_dialect)
                # Clean ANSI codes from error string
                error_str = str(e)
                error_str = re.sub(r'\x1b\[[0-9;]*m', '', error_str)
                error_str = re.sub(r'\033\[[0-9;]*m', '', error_str)
                error_str = re.sub(r'\[[0-9;]*m', '', error_str)
                error_details.append({
                    'statement': 'all',
                    'dialect': try_dialect or 'tsql',
                    'error': error_str,
                    'formatted': formatted_error
                })
            # Try next dialect if enabled
            if try_multiple_dialects:
                continue
            else:
                break  # Stop trying dialects
    
    # HYBRID FLOW: Try fallback if AST failed or extracted 0 columns
    # If AST succeeded but extracted 0 columns, or if parse error occurred, try fallback
    if (ast_succeeded and len(columns) == 0) or parse_error_occurred:
        fallback_columns = fallback_extract_columns(sql, enable_unqualified_resolution)
        
        if len(fallback_columns) > 0:
            # Fallback extracted columns - PARTIAL_OK (success)
            return fallback_columns, "PARTIAL_OK"
        else:
            # Fallback also extracted 0 columns
            if parse_error_occurred:
                # Parse error occurred and fallback failed
                if last_error:
                    formatted_error = format_parse_error(last_error, sql, last_dialect)
                    file_info = f"\nFile: {filepath}" if filepath else ""
                    print(f"\n{'='*80}", file=sys.stderr)
                    print(f"ERROR: Failed to parse SQL{file_info}", file=sys.stderr)
                    print(formatted_error, file=sys.stderr)
                    print(f"{'='*80}\n", file=sys.stderr)
                return [], "PARSE_ERROR"
            else:
                # AST succeeded but had zero columns and fallback also had zero
                return [], "ZERO_COLUMNS"
    
    # If we get here, something went wrong - return empty with error status
    return [], "PARSE_ERROR"


def process_sql_file(filepath: Path, dialect: Optional[str] = None, error_details: Optional[list] = None, enable_unqualified_resolution: bool = True, try_multiple_dialects: bool = False) -> tuple[list, str]:
    """
    Process a single SQL file and return all table.column references (all occurrences).
    
    Args:
        filepath: Path to SQL file
        dialect: SQL dialect to use (None defaults to 'tsql')
        error_details: Optional list to collect detailed error information
        enable_unqualified_resolution: Whether to infer table names for unqualified columns
        try_multiple_dialects: Whether to try multiple dialects on failure (default: False)
    
    Returns:
        tuple: (columns, status) where status is one of:
            - "SUCCESS": AST parsing succeeded and extracted columns
            - "PARTIAL_OK": Fallback parsing extracted columns
            - "PARSE_ERROR": ParseError occurred and fallback failed
            - "ZERO_COLUMNS": Both AST and fallback extracted zero columns
    """
    try:
        with open(filepath, "r", encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        if not content.strip():
            print(f"Warning: File {filepath} is empty", file=sys.stderr)
            return [], "ZERO_COLUMNS"
        
        # Preprocess SQL to handle edge cases
        content = preprocess_sql(content)
        
        if not content.strip():
            print(f"Warning: File {filepath} contains only comments/DDL (no query statements)", file=sys.stderr)
            return [], "ZERO_COLUMNS"
        
        return extract_table_columns(content, dialect, str(filepath), error_details, enable_unqualified_resolution, try_multiple_dialects)
    
    except FileNotFoundError:
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        return [], "PARSE_ERROR"
    except PermissionError:
        print(f"ERROR: Permission denied reading file: {filepath}", file=sys.stderr)
        return [], "PARSE_ERROR"
    except UnicodeDecodeError as e:
        print(f"ERROR: Unable to decode file {filepath}: {e}", file=sys.stderr)
        print("  Try checking file encoding or use a different encoding", file=sys.stderr)
        return [], "PARSE_ERROR"
    except Exception as e:
        print(f"ERROR: Unexpected error reading file {filepath}: {e}", file=sys.stderr)
        import traceback
        print(traceback.format_exc(), file=sys.stderr)
        return [], "PARSE_ERROR"


def setup_logging(log_file: Path):
    """Setup logging to both file and console."""
    # Create log file path
    log_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Extract table.column references from SQL files"
    )
    parser.add_argument(
        "files",
        nargs="*",
        help="SQL files or directories to process (default: all .sql files in current directory). Directories are searched recursively."
    )
    parser.add_argument(
        "--dialect",
        "-d",
        help="SQL dialect (tsql/mssql for SQL Server, postgres, mysql, snowflake, oracle, bigquery, etc.). Note: 'mssql' is automatically converted to 'tsql'."
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="Output file (default: columns.csv in current directory, use .xlsx for Excel)"
    )
    parser.add_argument(
        "--dataset",
        default="SQL_Parser",
        help="Dataset name to use in output (default: SQL_Parser). Note: Output format is now Filename, ColumnName"
    )
    parser.add_argument(
        "--no-unqualified-resolution",
        action="store_true",
        help="Disable inference of table names for unqualified columns"
    )
    parser.add_argument(
        "--try-multiple-dialects",
        action="store_true",
        help="Try multiple dialects on parse failure (default: only try specified dialect or tsql)"
    )
    
    args = parser.parse_args()
    
    # Normalize dialect if provided (convert mssql -> tsql, etc.)
    if args.dialect:
        args.dialect = normalize_dialect(args.dialect)
    
    # Determine output directory and file names
    if args.output:
        output_path = Path(args.output)
        # If output is a directory, use it; otherwise extract directory from file path
        if output_path.suffix:  # It's a file path
            output_dir = output_path.parent
            output_filename = output_path.name
        else:  # It's a directory
            output_dir = output_path
            output_filename = "columns.csv"
    else:
        # Default to output/ directory in current directory
        output_dir = Path("output")
        output_filename = "columns.csv"
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create Error_Reports subdirectory
    error_reports_dir = output_dir / "Error_Reports"
    error_reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Set up file paths
    output_path = output_dir / output_filename
    log_file = output_dir / output_path.with_suffix('.log').name
    errors_file = output_dir / "errors.txt"
    
    # Setup logging (log file will be in output directory)
    logger = setup_logging(log_file)
    
    logger.info("="*80)
    logger.info("SQL Column Extractor - Starting")
    logger.info("="*80)
    logger.info(f"Output file: {output_path.absolute()}")
    logger.info(f"Log file: {log_file.absolute()}")
    
    # Determine which files to process
    sql_files = []
    
    # Check if FOLDER_PATH is set (highest priority)
    if FOLDER_PATH and FOLDER_PATH.strip():
        folder_path = Path(FOLDER_PATH.strip())
        if not folder_path.exists():
            error_msg = f"FOLDER_PATH not found: {folder_path}"
            logger.error(error_msg)
            logger.error("Please check the FOLDER_PATH variable at the top of the script.")
            print(f"Error: {error_msg}", file=sys.stderr)
            print("Please check the FOLDER_PATH variable at the top of the script.", file=sys.stderr)
            return
        
        if folder_path.is_file():
            # Single file
            if folder_path.suffix.lower() == '.sql':
                sql_files.append(folder_path)
                logger.info(f"Using FOLDER_PATH (file): {folder_path}")
                print(f"Using FOLDER_PATH: {folder_path}")
            else:
                error_msg = f"FOLDER_PATH is not a SQL file: {folder_path}"
                logger.error(error_msg)
                print(f"Error: {error_msg}", file=sys.stderr)
                return
        elif folder_path.is_dir():
            # Directory - find all SQL files recursively
            found_files = sorted(folder_path.rglob("*.sql"))
            sql_files.extend(found_files)
            logger.info(f"Using FOLDER_PATH (directory): {folder_path}")
            logger.info(f"Found {len(found_files)} SQL file(s)")
            print(f"Using FOLDER_PATH: {folder_path}")
            print(f"Found {len(found_files)} SQL file(s) in {folder_path}")
        else:
            error_msg = f"FOLDER_PATH is not a valid file or directory: {folder_path}"
            logger.error(error_msg)
            print(f"Error: {error_msg}", file=sys.stderr)
            return
    
    elif args.files:
        # Process provided files/directories from command-line
        for file_path in args.files:
            path = Path(file_path)
            if not path.exists():
                warning_msg = f"Path not found: {path}"
                logger.warning(warning_msg)
                print(f"Warning: {warning_msg}", file=sys.stderr)
                continue
            
            if path.is_file():
                # Single file
                if path.suffix.lower() == '.sql':
                    sql_files.append(path)
                else:
                    warning_msg = f"Skipping non-SQL file: {path}"
                    logger.warning(warning_msg)
                    print(f"Warning: {warning_msg}", file=sys.stderr)
            elif path.is_dir():
                # Directory - find all SQL files recursively
                found_files = sorted(path.rglob("*.sql"))
                sql_files.extend(found_files)
                logger.info(f"Found {len(found_files)} SQL file(s) in {path}")
                print(f"Found {len(found_files)} SQL file(s) in {path}")
    else:
        # Default to current directory (recursive)
        sql_files = sorted(Path.cwd().rglob("*.sql"))
        logger.info(f"Using current directory: {Path.cwd()}")
    
    if not sql_files:
        error_msg = "No SQL files found to process."
        logger.error(error_msg)
        print(error_msg)
        return
    
    # Collect all column references with file tracking
    # Structure: list of tuples (filename, column_name)
    # We'll deduplicate per file (keep unique columns per file)
    column_data = []
    file_columns = {}
    
    # Error tracking
    files_with_errors = []  # Files that threw exceptions
    files_with_zero_columns = []  # Files that processed but found 0 columns
    files_successful = []  # Files that successfully found columns
    
    logger.info(f"\nProcessing {len(sql_files)} SQL file(s)...")
    print(f"\nProcessing {len(sql_files)} SQL file(s)...")
    
    for sql_file in sql_files:
        logger.info(f"Processing: {sql_file}")
        print(f"  Processing: {sql_file}")
        file_error_details = []  # Collect error details for this file
        try:
            # Determine flags from command-line arguments
            enable_unqualified = not args.no_unqualified_resolution
            try_multiple = args.try_multiple_dialects
            
            # Process SQL file - returns (columns, status) tuple
            result = process_sql_file(
                sql_file, 
                args.dialect, 
                file_error_details,
                enable_unqualified_resolution=enable_unqualified,
                try_multiple_dialects=try_multiple
            )
            
            # Safely unpack the result tuple
            if not isinstance(result, tuple) or len(result) != 2:
                raise ValueError(f"process_sql_file returned unexpected result: {result} (expected tuple of length 2)")
            
            columns, status = result
            
            # Ensure status is a valid string
            if not isinstance(status, str):
                logger.warning(f"Invalid status type from process_sql_file: {type(status)}, defaulting to PARSE_ERROR")
                status = "PARSE_ERROR"
            
            # Parse filename to get report name and dataset
            report_name, dataset = parse_filename(sql_file)
            logger.info(f"  Parsed: Report='{report_name}', Dataset='{dataset}'")
            
            # Skip files with excluded dataset names
            if should_skip_dataset(dataset):
                logger.info(f"  Skipping file (dataset '{dataset}' matches exclusion pattern)")
                print(f"  Skipping: {sql_file.name} (dataset '{dataset}' matches exclusion pattern)")
                continue
            
            # Store full relative path or just filename for file_columns tracking
            file_key = str(sql_file.relative_to(Path.cwd())) if sql_file.is_relative_to(Path.cwd()) else str(sql_file)
            
            # Filter out wildcard columns (table.*) and deduplicate per file
            # Keep only unique columns per file (can have duplicates across files)
            unique_columns_for_file = []
            seen_in_file = set()
            wildcard_count = 0
            
            for col in columns:
                # Skip wildcard columns (table.*)
                if col.endswith('.*'):
                    wildcard_count += 1
                    logger.debug(f"Skipping wildcard column: {col}")
                    continue
                # Deduplicate within this file
                if col not in seen_in_file:
                    seen_in_file.add(col)
                    unique_columns_for_file.append(col)
            
            file_columns[file_key] = unique_columns_for_file
            
            # Handle different statuses
            if status == "SUCCESS":
                # AST parsing succeeded and extracted columns
                files_successful.append(file_key)
                for col in unique_columns_for_file:
                    column_data.append((report_name, dataset, col))
                logger.info(f"  Found {len(unique_columns_for_file)} unique columns in {report_name} ({dataset}) [AST parsing]")
            elif status == "PARTIAL_OK":
                # Fallback parsing extracted columns - treat as success
                files_successful.append(file_key)
                for col in unique_columns_for_file:
                    column_data.append((report_name, dataset, col))
                logger.info(f"  Found {len(unique_columns_for_file)} unique columns in {report_name} ({dataset}) [Fallback parsing - PARTIAL_OK]")
            elif status == "PARSE_ERROR":
                # Hard parse error - copy to Error_Reports
                import shutil
                # Build error message from parse errors if available
                error_msg = "Parse error occurred"
                if file_error_details:
                    error_parts = []
                    for parse_err in file_error_details:
                        err_text = parse_err.get('error', 'Unknown parse error')
                        dialect = parse_err.get('dialect', 'unknown')
                        stmt = parse_err.get('statement', 'unknown')
                        error_parts.append(f"Statement {stmt} (dialect: {dialect}): {err_text}")
                    if error_parts:
                        error_msg = "; ".join(error_parts[:3])  # Limit to first 3 errors
                        if len(error_parts) > 3:
                            error_msg += f" ... and {len(error_parts) - 3} more"
                
                error_info = {
                    'file': file_key,
                    'report_name': report_name,
                    'dataset': dataset,
                    'status': status,
                    'error': error_msg,  # Add error field for consistency
                    'total_extracted': len(columns),
                    'wildcards_filtered': wildcard_count
                }
                if file_error_details:
                    error_info['parse_errors'] = file_error_details
                    logger.warning(f"  Parse error in {report_name} ({dataset}) - Parse errors detected: {len(file_error_details)}")
                    for parse_err in file_error_details:
                        logger.warning(f"    Parse error: Statement {parse_err.get('statement', 'unknown')}, Dialect: {parse_err.get('dialect', 'unknown')}")
                files_with_errors.append(error_info)
                
                # Copy file with parse error to Error_Reports directory
                try:
                    error_report_path = error_reports_dir / sql_file.name
                    shutil.copy2(sql_file, error_report_path)
                    logger.info(f"Copied file with parse error to Error_Reports: {error_report_path}")
                except Exception as copy_error:
                    logger.error(f"Failed to copy file {sql_file} to Error_Reports: {copy_error}", exc_info=True)
            elif status == "ZERO_COLUMNS":
                # Zero columns - don't copy to Error_Reports (not a hard error)
                zero_col_info = {
                    'file': file_key,
                    'report_name': report_name,
                    'dataset': dataset,
                    'status': status,
                    'total_extracted': len(columns),
                    'wildcards_filtered': wildcard_count
                }
                if file_error_details:
                    zero_col_info['parse_errors'] = file_error_details
                    logger.warning(f"  No columns found in {report_name} ({dataset}) - Parse errors detected: {len(file_error_details)}")
                else:
                    logger.warning(f"  No columns found in {report_name} ({dataset}) - Total extracted: {len(columns)}, Wildcards filtered: {wildcard_count}")
                files_with_zero_columns.append(zero_col_info)
        except Exception as e:
            import traceback
            error_msg = f"Error processing {sql_file}: {e}"
            full_traceback = traceback.format_exc()
            logger.error(error_msg, exc_info=True)
            logger.error(f"Full traceback for {sql_file}:\n{full_traceback}")
            print(f"  {error_msg}", file=sys.stderr)
            
            # Copy failed file to Error_Reports directory
            try:
                error_report_path = error_reports_dir / sql_file.name
                shutil.copy2(sql_file, error_report_path)
                logger.info(f"Copied failed file to Error_Reports: {error_report_path}")
            except Exception as copy_error:
                logger.error(f"Failed to copy error file {sql_file} to Error_Reports: {copy_error}", exc_info=True)
            
            # Combine error details if available
            error_info = {
                'file': str(sql_file),
                'error': str(e),
                'traceback': full_traceback
            }
            
            # Add detailed parse errors if available
            if file_error_details:
                error_info['parse_errors'] = file_error_details
            
            files_with_errors.append(error_info)
            continue
    
    # Count unique vs total
    all_columns = [col for _, _, col in column_data]
    unique_columns = sorted(set(all_columns))
    total_count = len(column_data)
    unique_count = len(unique_columns)
    
    logger.info(f"\nFound {total_count} total table.column references ({unique_count} unique)")
    logger.info(f"Across {len(file_columns)} file(s)")
    print(f"\nFound {total_count} total table.column references ({unique_count} unique)")
    print(f"Across {len(file_columns)} file(s)")
    
    # Output to CSV or Excel
    # Track if CSV was written (for safety check)
    csv_written = False
    
    if output_path.suffix.lower() == '.xlsx':
        # Excel output
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            # Sort column_data by ReportName, Dataset, ColumnName
            sorted_column_data = sorted(column_data, key=lambda x: (
                str(x[0]) if x[0] else '',  # ReportName
                str(x[1]) if x[1] else '',  # Dataset
                str(x[2]) if x[2] else ''   # ColumnName
            ))
            
            # Create DataFrame with ReportName, Dataset, and ColumnName
            # Ensure all values are strings and handle None/empty values
            df = pd.DataFrame({
                'ReportName': [str(report) if report else '' for report, _, _ in sorted_column_data],
                'Dataset': [str(dataset) if dataset else '' for _, dataset, _ in sorted_column_data],
                'ColumnName': [str(col) if col else '' for _, _, col in sorted_column_data]
            })
            
            df.to_excel(output_path, index=False)
            
            # Auto-format column widths and add filters
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add filter to first row
            ws.auto_filter.ref = ws.dimensions
            
            wb.save(output_path)
            
            abs_path = output_path.absolute()
            logger.info(f"Excel output written successfully: {abs_path}")
            logger.info(f"  Rows: {len(df)} (unique columns per file)")
            logger.info(f"  Columns: ReportName, Dataset, ColumnName")
            logger.info(f"  Sorted by: ReportName, Dataset, ColumnName")
            logger.info(f"  Auto-formatted column widths and filters applied")
            print("\n" + "="*80)
            print("OUTPUT FILE LOCATION:")
            print("="*80)
            print(f"  {abs_path}")
            print("="*80)
            print(f"\n✓ Output written to: {abs_path}")
            print(f"  Rows: {len(df)} (unique columns per file)")
            print(f"  Columns: ReportName, Dataset, ColumnName")
            print(f"  Sorted by: ReportName, Dataset, ColumnName")
            print(f"  Auto-formatted column widths and filters applied")
            print(f"  Output directory: {output_dir.absolute()}")
            print(f"  Log file: {log_file.absolute()}")
            print(f"  Errors file: {errors_file.absolute()}")
            print(f"  Error reports: {error_reports_dir.absolute()}")
            csv_written = True  # Excel written successfully
            
        except ImportError:
            error_msg = "pandas and openpyxl required for Excel output"
            logger.warning(error_msg)
            logger.warning("Falling back to CSV output...")
            print("\nError: pandas and openpyxl required for Excel output.")
            print("Install with: pip install pandas openpyxl")
            print("\nFalling back to CSV output...")
            output_path = output_path.with_suffix('.csv')
            # Update log file path if we changed output path
            log_file = output_path.with_suffix('.log')
    
    if output_path.suffix.lower() == '.csv':
        # CSV output
        try:
            # Sort column_data by ReportName, Dataset, ColumnName
            sorted_column_data = sorted(column_data, key=lambda x: (
                str(x[0]) if x[0] else '',  # ReportName
                str(x[1]) if x[1] else '',  # Dataset
                str(x[2]) if x[2] else ''   # ColumnName
            ))
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['ReportName', 'Dataset', 'ColumnName'])  # Header
                
                for report_name, dataset, col in sorted_column_data:
                    # Ensure all values are strings and handle None/empty values
                    safe_report = str(report_name) if report_name else ''
                    safe_dataset = str(dataset) if dataset else ''
                    safe_col = str(col) if col else ''
                    writer.writerow([safe_report, safe_dataset, safe_col])
            
            abs_path = output_path.absolute()
            logger.info(f"CSV output written successfully: {abs_path}")
            logger.info(f"  Rows: {len(sorted_column_data)} (unique columns per file)")
            logger.info(f"  Columns: ReportName, Dataset, ColumnName")
            logger.info(f"  Sorted by: ReportName, Dataset, ColumnName")
            print("\n" + "="*80)
            print("OUTPUT FILE LOCATION:")
            print("="*80)
            print(f"  {abs_path}")
            print("="*80)
            print(f"\n✓ Output written to: {abs_path}")
            print(f"  Rows: {len(sorted_column_data)} (unique columns per file)")
            print(f"  Columns: ReportName, Dataset, ColumnName")
            print(f"  Sorted by: ReportName, Dataset, ColumnName")
            print(f"  Output directory: {output_dir.absolute()}")
            print(f"  Log file: {log_file.absolute()}")
            print(f"  Errors file: {errors_file.absolute()}")
            print(f"  Error reports: {error_reports_dir.absolute()}")
            csv_written = True  # CSV written successfully
        except Exception as e:
            error_msg = f"Failed to write CSV file {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            print(f"\nERROR: {error_msg}", file=sys.stderr)
            import traceback
            print(traceback.format_exc(), file=sys.stderr)
    
    # Ensure CSV is written if it wasn't written yet (shouldn't happen, but safety check)
    if not csv_written and output_path.suffix.lower() != '.xlsx':
        logger.warning(f"CSV file was not written. Output path: {output_path}")
        print(f"\nWARNING: CSV file was not written. Output path: {output_path}", file=sys.stderr)
    
    # Show summary by file
    logger.info("\n" + "="*60)
    logger.info("SUMMARY BY FILE")
    logger.info("="*60)
    print("\n" + "="*60)
    print("SUMMARY BY FILE")
    print("="*60)
    for file_key, cols in sorted(file_columns.items()):
        # Parse filename for display
        file_path = Path(file_key)
        report_name, dataset = parse_filename(file_path)
        unique_in_file = len(set(cols))
        summary_line = f"  {report_name} ({dataset}): {len(cols)} unique columns"
        logger.info(summary_line)
        print(summary_line)
    
    # Show error summary
    print("\n" + "="*60)
    print("PROCESSING SUMMARY")
    print("="*60)
    total_files = len(sql_files)
    successful_count = len(files_successful)
    zero_columns_count = len(files_with_zero_columns)
    error_count = len(files_with_errors)
    
    logger.info("\n" + "="*60)
    logger.info("PROCESSING SUMMARY")
    logger.info("="*60)
    logger.info(f"Total files processed: {total_files}")
    logger.info(f"Successfully processed with columns: {successful_count}")
    logger.info(f"Files with 0 columns found: {zero_columns_count}")
    logger.info(f"Files with errors: {error_count}")
    
    print(f"Total files processed: {total_files}")
    print(f"Successfully processed with columns: {successful_count}")
    print(f"Files with 0 columns found: {zero_columns_count}")
    print(f"Files with errors: {error_count}")
    
    # Show files with zero columns
    if files_with_zero_columns:
        print(f"\nFiles with 0 columns found ({zero_columns_count}):")
        logger.info(f"\nFiles with 0 columns found ({zero_columns_count}):")
        for file_info in files_with_zero_columns[:20]:  # Show first 20
            msg = f"  {file_info['report_name']} ({file_info['dataset']}): {file_info['file']}"
            if file_info['total_extracted'] > 0:
                msg += f" - {file_info['total_extracted']} columns extracted but filtered (wildcards: {file_info['wildcards_filtered']})"
            else:
                msg += " - No columns extracted (may be DDL-only or parse error)"
            logger.info(msg)
            print(msg)
        if len(files_with_zero_columns) > 20:
            remaining = len(files_with_zero_columns) - 20
            msg = f"  ... and {remaining} more files with 0 columns"
            logger.info(msg)
            print(msg)
    
    # Show files with errors
    if files_with_errors:
        print(f"\nFiles with processing errors ({error_count}):")
        logger.error(f"\nFiles with processing errors ({error_count}):")
        for file_info in files_with_errors[:20]:  # Show first 20
            error_msg = file_info.get('error', 'Unknown error (see parse_errors for details)')
            msg = f"  {file_info['file']}: {error_msg}"
            logger.error(msg)
            print(msg)
            # Log full traceback to log file
            if 'traceback' in file_info:
                logger.error(f"Full traceback for {file_info['file']}:\n{file_info['traceback']}")
        if len(files_with_errors) > 20:
            remaining = len(files_with_errors) - 20
            msg = f"  ... and {remaining} more files with errors"
            logger.info(msg)
            print(msg)
            # Log tracebacks for remaining files
            for file_info in files_with_errors[20:]:
                error_msg = file_info.get('error', 'Unknown error (see parse_errors for details)')
                logger.error(f"Error in {file_info['file']}: {error_msg}")
                if 'traceback' in file_info:
                    logger.error(f"Full traceback for {file_info['file']}:\n{file_info['traceback']}")
    
    # Write errors to errors.txt file (always write if there are any issues)
    if files_with_errors or files_with_zero_columns:
        try:
            with open(errors_file, 'w', encoding='utf-8') as f:
                f.write("="*80 + "\n")
                f.write("ERROR REPORT\n")
                f.write("="*80 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total files processed: {len(sql_files)}\n")
                f.write(f"Files with errors: {len(files_with_errors)}\n")
                f.write(f"Files with 0 columns: {len(files_with_zero_columns)}\n")
                f.write(f"Files successfully processed: {len(files_successful)}\n\n")
                
                if files_with_errors:
                    f.write(f"FILES WITH PROCESSING ERRORS ({len(files_with_errors)}):\n")
                    f.write("-"*80 + "\n")
                    for file_info in files_with_errors:
                        f.write(f"\nFile: {file_info['file']}\n")
                        error_msg = file_info.get('error', 'Unknown error (see parse_errors for details)')
                        f.write(f"Error: {error_msg}\n")
                        
                        # Write detailed parse errors if available
                        if 'parse_errors' in file_info and file_info['parse_errors']:
                            f.write(f"\nDetailed Parse Errors:\n")
                            for parse_error in file_info['parse_errors']:
                                f.write(f"\n  Statement #{parse_error.get('statement', 'unknown')}\n")
                                f.write(f"  Dialect: {parse_error.get('dialect', 'unknown')}\n")
                                if 'dialects_tried' in parse_error:
                                    f.write(f"  Dialects tried: {parse_error['dialects_tried']}\n")
                                f.write(f"  Error: {parse_error.get('error', 'Unknown error')}\n")
                                f.write(f"\n  Detailed Error Information:\n")
                                # Indent the formatted error
                                formatted = parse_error.get('formatted', '')
                                for line in formatted.split('\n'):
                                    f.write(f"    {line}\n")
                        
                        if 'traceback' in file_info:
                            f.write(f"\nTraceback:\n{file_info['traceback']}\n")
                        f.write("-"*80 + "\n")
                    f.write("\n")
                
                if files_with_zero_columns:
                    f.write(f"FILES WITH 0 COLUMNS FOUND ({len(files_with_zero_columns)}):\n")
                    f.write("-"*80 + "\n")
                    for file_info in files_with_zero_columns:
                        f.write(f"\nFile: {file_info['file']}\n")
                        f.write(f"Report: {file_info['report_name']} ({file_info['dataset']})\n")
                        if file_info['total_extracted'] > 0:
                            f.write(f"Reason: {file_info['total_extracted']} columns extracted but filtered ")
                            f.write(f"(wildcards: {file_info['wildcards_filtered']})\n")
                        else:
                            # Check if this is a parse error or just DDL/empty
                            if 'parse_errors' in file_info and file_info['parse_errors']:
                                f.write(f"Reason: Parse error(s) prevented column extraction\n")
                            else:
                                f.write(f"Reason: No columns extracted (may be DDL-only, empty file, or parse error)\n")
                        
                        # Write detailed parse errors if available
                        if 'parse_errors' in file_info and file_info['parse_errors']:
                            f.write(f"\nDetailed Parse Errors:\n")
                            for parse_error in file_info['parse_errors']:
                                f.write(f"\n  Statement #{parse_error.get('statement', 'unknown')}\n")
                                f.write(f"  Dialect: {parse_error.get('dialect', 'unknown')}\n")
                                if 'dialects_tried' in parse_error:
                                    f.write(f"  Dialects tried: {parse_error['dialects_tried']}\n")
                                f.write(f"  Error: {parse_error.get('error', 'Unknown error')}\n")
                                f.write(f"\n  Detailed Error Information:\n")
                                # Indent the formatted error
                                formatted = parse_error.get('formatted', '')
                                for line in formatted.split('\n'):
                                    f.write(f"    {line}\n")
                        
                        f.write("-"*80 + "\n")
                    f.write("\n")
                
                f.write(f"\nTotal files with errors: {len(files_with_errors)}\n")
                f.write(f"Total files with 0 columns: {len(files_with_zero_columns)}\n")
                f.write(f"\nAll error files have been copied to: {error_reports_dir.absolute()}\n")
            
            logger.info(f"Error report written to: {errors_file.absolute()}")
            print(f"\nError report written to: {errors_file.absolute()}")
        except Exception as e:
            logger.error(f"Failed to write error report file: {e}", exc_info=True)
    
    # Show sample of extracted columns
    logger.info("\n" + "="*60)
    logger.info("SAMPLE OUTPUT (first 20 rows)")
    logger.info("="*60)
    print("\n" + "="*60)
    print("SAMPLE OUTPUT (first 20 rows)")
    print("="*60)
    for report_name, dataset, col in column_data[:20]:
        sample_line = f"  {report_name} | {dataset} | {col}"
        logger.info(sample_line)
        print(sample_line)
    if len(column_data) > 20:
        remaining_msg = f"  ... and {len(column_data) - 20} more rows"
        logger.info(remaining_msg)
        print(remaining_msg)
    
    logger.info("="*80)
    logger.info("SQL Column Extractor - Completed Successfully")
    logger.info("="*80)


if __name__ == "__main__":
    main()

